from concurrent.futures import wait, as_completed, ThreadPoolExecutor
import requests
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import re
import time
import random

codepat = re.compile(r'(OUTCODE|REGION)\^(\d+)')

Available_outcode = dict()
Unavilable_outcode = set()

url_prefix = 'https://www.rightmove.co.uk/property-to-rent/search.html?searchLocation='


def extract_code_from_excel():
    start_time = time.time()
    outcode_list = []
    wb = load_workbook('postcode-outcodes.xlsx')
    ws = wb['postcode-outcodes']
    for row in range(2, ws.max_row + 1):
        outcode = ws['B' + str(row)].value
        outcode_list.append(outcode)
    print("Extract from excel spend:", time.time()-start_time)
    return outcode_list

def get_code_from_rightmove(outcode):
    sleep_time = random.random()
    time.sleep(sleep_time)
    # print("%r is in the func"% outcode)
    url = url_prefix + outcode
    resp = requests.get(url)
    soup = BeautifulSoup(resp.text, "html.parser")
    find_error = soup.find("div", class_='errorbox')
    if find_error:
        Unavilable_outcode.add(outcode)
    else:
        get_value = soup.find('input', {'id': 'locationIdentifier'}).get('value')
        get_code = codepat.match(get_value)
        if get_code:
            Available_outcode[outcode] = (get_code.group(1),get_code.group(2))
        else:
            Unavilable_outcode.add(outcode)
            # print("error:", get_code)
    print("%s is finished, sleepint time is %s:" % (outcode, sleep_time))

def save_code_to_excel(title_name, l):
    title = title_name
    wb = load_workbook('postcode-outcodes.xlsx')
    source = wb.active
    ws = wb.copy_worksheet(source) if title not in wb.sheetnames else wb[title]
    ws.title = title
    for row in range(2, ws.max_row + 1):
        outcode = ws['B' + str(row)].value
        if outcode in l:
            ws['E' + str(row)] = 1
            ws['F' + str(row)] = l[outcode][0]
            ws['G' + str(row)] = l[outcode][1]
        else:
            ws['E' + str(row)] = 0
    wb.save('postcode-outcodes.xlsx')
 
def run_task(task):
    with ThreadPoolExecutor(max_workers=5) as executor:
        future_to_url = { executor.submit(get_code_from_rightmove, outcode): outcode for outcode in task}
        # wait_futures = wait(future_to_url)
        # print(wait_futures)
        for future in as_completed(future_to_url):
            url = future_to_url[future]
            try:
                data = future.result()
            except Exception as exc:
                print('%r generated an exception: %s' % (url, exc))
            # else:
            #     print('%r is completed' % url)
 
def fake_result():
    outcode_list = extract_code_from_excel()
    result = {}
    for outcode in outcode_list:
        result[outcode] = random.choice([0,1,2])
    return result

def test_code_from_rightmove(outcode):
    url = url_prefix + outcode
    resp = requests.get(url)
    print(resp.status_code)
    soup = BeautifulSoup(resp.text, "html.parser")
    find_error = soup.find("div", class_='errorbox')
    if find_error:
        print(url, "result:","error")
    else:
        get_value = soup.find('input', {'id': 'locationIdentifier'}).get('value')
        # print(get_value)
        get_code = codepat.match(get_value)
        if get_code:
            Available_outcode[outcode] = get_code.group(1)
            # print(get_value,get_code.group(1))
        else:
            Unavilable_outcode.add(outcode)
            # print("error:", get_code)
    print("%s is finished" % outcode)

if __name__ == "__main__":
    start_time = time.time()
    # fake_list =fake_result()
    # print(fake_list)
    # test_code_from_rightmove('WA14')
    outcode_list = extract_code_from_excel()
    run_task(outcode_list)
    save_code_to_excel("ran_test",Available_outcode)
    print("Unavilable_outcode: ", Unavilable_outcode, len(Unavilable_outcode))
    print("Successful outcode number: ", len(Available_outcode))
    print("Spend time: ", time.time()-start_time)

