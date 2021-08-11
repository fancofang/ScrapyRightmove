import scrapy
from urllib.parse import urlparse, parse_qs, urlencode
from scrapyrightmove.items import ScrapyrightmoveItem
from scrapyrightmove.utility import *
from openpyxl import load_workbook
import time
import logging


class Myspider(scrapy.Spider):
    name = 'rightmove_front'
    seen = set()
    count = 0
    
    # def __init__(self, outcode):
    #     self.outcode_list = outcode

    # set when finished call back spider_closed func
    @classmethod
    def from_crawler(cls, crawler, *args, **kwargs):
        spider = super(Myspider, cls).from_crawler(crawler, *args, **kwargs)
        crawler.signals.connect(spider.spider_closed, signal=scrapy.signals.spider_closed)
        return spider

    def spider_closed(self, spider):
        print('Spider closed: %s, total properties: %s' % (spider.name, self.count))

    def extract_code_from_excel(self):
        import os
        path = os.getcwd()
        excel_path = os.path.join(path, 'postcode', 'postcode-outcodes1.xlsx')
        outcode_list = []
        wb = load_workbook(excel_path)
        ws = wb['ran']
        for row in range(2, ws.max_row + 1):
            if ws['E' + str(row)].value != 0 and ws['F' + str(row)].value == 'OUTCODE':
                outcode_list.append( ws['G' + str(row)].value)
        return outcode_list

    def start_requests(self):
        url_prefix = 'https://www.rightmove.co.uk/property-to-rent/find.html?locationIdentifier=OUTCODE%5E'
        outcode_list = self.extract_code_from_excel()
        
        for outcode in outcode_list:
            url = url_prefix + str(outcode)
            yield scrapy.Request(url=url, callback=self.parse)
    
    
    
    # start_urls = [
    #     'https://www.rightmove.co.uk/property-to-rent/find.html?locationIdentifier=REGION%5E93917'
    #     '&maxPrice=800&includeLetAgreed=false',
    #
    #     'https://www.rightmove.co.uk/property-to-rent/find.html?locationIdentifier=REGION%5E93917'
    #     '&maxPrice=1000&minPrice=800&includeLetAgreed=false',
    #
    #     'https://www.rightmove.co.uk/property-to-rent/find.html?locationIdentifier=REGION%5E93917'
    #     '&maxPrice=1200&minPrice=1000&includeLetAgreed=false',
    #
    #     'https://www.rightmove.co.uk/property-to-rent/find.html?locationIdentifier=REGION%5E93917'
    #     '&maxPrice=1300&minPrice=1200&includeLetAgreed=false',
    #
    #     'https://www.rightmove.co.uk/property-to-rent/find.html?locationIdentifier=REGION%5E93917'
    #     '&maxPrice=1400&minPrice=1300&includeLetAgreed=false',
    #
    #     'https://www.rightmove.co.uk/property-to-rent/find.html?locationIdentifier=REGION%5E93917'
    #     '&maxPrice=1500&minPrice=1400&sortType=2&includeLetAgreed=false',
    #
    #     'https://www.rightmove.co.uk/property-to-rent/find.html?locationIdentifier=REGION%5E93917'
    #     '&maxPrice=1750&minPrice=1500&sortType=2&includeLetAgreed=false',
    #
    #     'https://www.rightmove.co.uk/property-to-rent/find.html?locationIdentifier=REGION%5E93917'
    #     '&maxPrice=2000&minPrice=1750&sortType=2&includeLetAgreed=false',
    #
    #     'https://www.rightmove.co.uk/property-to-rent/find.html?locationIdentifier=REGION%5E93917'
    #     '&maxPrice=2500&minPrice=2000&sortType=2&includeLetAgreed=false',
    #
    #     'https://www.rightmove.co.uk/property-to-rent/find.html?locationIdentifier=REGION%5E93917'
    #     '&maxPrice=3000&minPrice=2500&sortType=2&includeLetAgreed=false',
    #
    #     'https://www.rightmove.co.uk/property-to-rent/find.html?locationIdentifier=REGION%5E93917'
    #     '&maxPrice=3500&minPrice=3000&sortType=2&includeLetAgreed=false',
    #
    #     'https://www.rightmove.co.uk/property-to-rent/find.html?locationIdentifier=REGION%5E93917'
    #     '&maxPrice=5000&minPrice=3500&sortType=2&includeLetAgreed=false',
    #
    #     'https://www.rightmove.co.uk/property-to-rent/find.html?locationIdentifier=REGION%5E93917'
    #     '&maxPrice=20000&minPrice=5000&sortType=2&includeLetAgreed=false'
    #
    #     'https://www.rightmove.co.uk/property-to-rent/find.html?locationIdentifier=REGION%5E93917'
    #     '&minPrice=20000&sortType=2&includeLetAgreed=false'
    # ]
    
    
    
    # def start_requests(self):
    #     url_prefix = 'https://www.rightmove.co.uk/property-for-sale/find.html?locationIdentifier=OUTCODE%5E412'
    
    def parse(self, response):
        self.logger.info('Parse function called on %s', response.url)
        Extract_total = response.xpath('//span[@class="searchHeader-resultCount"]/text()').get()
        Extract_total = Extract_total.replace(',','')
        This_area_total = int(Extract_total)

        parsed = urlparse(response.url)
        querys = parse_qs(parsed.query)
        para = {k: v[0] for k, v in querys.items()}
        para['index'] = int(para['index']) + 24 if 'index' in para else 24
        if para['index'] < This_area_total:
            print("地区：%s; 总房源数: %s；目前第%s页；准备下一页爬虫" % (para['locationIdentifier'],This_area_total, para['index']//24))
            data = urlencode(para)
            next_page = 'https://www.rightmove.co.uk/property-to-rent/find.html?' + data
            yield response.follow(next_page, callback=self.parse)
            
        htmls = response.xpath('//a[contains(@href,"/properties")]/@href').getall()
        for i in htmls:
            if i not in self.seen:
                self.seen.add(i)
                yield response.follow(i, callback=self.parse_details)

        
        
        ## invalid extracted url way
        # htmls = response.xpath('//a[contains(@href,"/property-to-rent/property")]/@href').getall()
        # htmls = response.xpath('//a[contains(@href,"/properties")]/@href').getall()
        # for i in htmls:
        #     if i not in self.seen:
        #         self.seen.add(i)
        #         yield response.follow(i, callback=self.parse_details)
        # parsed = urlparse(response.url)
        # querys = parse_qs(parsed.query)
        # para = {k: v[0] for k, v in querys.items()}
        # if 'index' not in para:
        #     para['index'] = '24'
        # else:
        #     para['index'] = str(int(para['index']) + 24)
        # data = urlencode(para)
        # next_page = 'https://www.rightmove.co.uk/property-to-rent/find.html?' + data
        # num = response.xpath('//span[@class="searchHeader-resultCount"]/text()').get()
        # num = num.replace(',', '')
        # print("总共房源数:", num)
        # print('目前index:', para['index'])
        #
        # if int(para['index']) < int(num):
        #     print("准备下一页爬虫")
        #     yield response.follow(next_page, callback=self.parse)
        # else:
        #     raise CloseSpider('完成爬虫,退出!')

    def parse_details(self, response):
        self.count += 1
        print("The crawling page is %s." % response.url)
        item = ScrapyrightmoveItem()
        try:
            data = extract_details(response)
            item['id'] = extract_id(data)
            item['title'] = extract_title(data)
            item['address'] = extract_address(data)
            item['agent'] = extract_agentdetails(data)
            item['rent'] = extract_rent(data)
            item['latitude'], item['longitude'] = extract_location(data)
            item['postcode'] = extract_postcode(data)
            # due to original url are filtered by letagree=False, so don't need to extract
            # letagreed = extract_letagreed(data)
            item['create_date'] = extract_addtime(data)
            item['url'] = response.url
            item['let_agreed'] = extract_letagreed(data)
            item['thumbnail'] = extract_image(data)
            # item = ScrapyrightmoveItem(id=id, url=response.url, agent=agent,
            #                            addTime=addtime, title=title,
            #                            address=address, rent={'date': date.today(), 'price': rent},
            #                            letAgreed=letagreed, postcode=postcode, location=location
            #                            )
            yield item
            # print(item)
        except Exception as e:
            self.logger.error("parse_details error: %s" % str(e))
            print("extraction error:", response.url)


        # data = extract_details(response)
        # try:
        #     id = extract_id(data)
        #     title = extract_title(data)
        #     address = extract_address(data)
        #     agent = extract_agentdetails(data)
        #     rent = extract_rent(data)
        #     location = extract_location(data)
        #     postcode = extract_postcode(data)
        #     # due to original url are filtered by letagree=False, so don't need to extract
        #     # letagreed = extract_letagreed(data)
        #     letagreed = False
        #     addtime = extract_addtime(data)
        #
        #     item = ScrapyrightmoveItem(id=id, url=response.url, agent=agent,
        #                                addTime=addtime, title=title,
        #                                address=address, rent={'date': date.today(), 'price': rent},
        #                                letAgreed=letagreed, postcode=postcode, location=location
        #                                )
        #     yield item
        #     # print(item)
        # except Exception as e:
        #     self.logger.error(str(e))

