import scrapy
from urllib.parse import urlparse, parse_qs, urlencode
from scrapyrightmove.items import ScrapyrightmoveItem
from scrapyrightmove.utility import *
from openpyxl import load_workbook
from collections import defaultdict

class Myspider(scrapy.Spider):
    name = 'rightmove_spider9'
    seen = set()
    count = 0
    try_again = defaultdict(int)
    
    # def __init__(self, outcode):
    #     self.outcode_list = outcode

    # set when finished call back spider_closed func
    @classmethod
    def from_crawler(cls, crawler, *args, **kwargs):
        spider = super(Myspider, cls).from_crawler(crawler, *args, **kwargs)
        crawler.signals.connect(spider.spider_closed, signal=scrapy.signals.spider_closed)
        return spider

    def spider_closed(self, spider):
        print('Spider closed: %s, total properties: %s' % (spider.name, self.count))

    def extract_code_from_excel(self):
        import os
        path = os.getcwd()
        excel_path = os.path.join(path, 'postcode', 'postcode-outcodes9.xlsx')
        outcode_list = []
        wb = load_workbook(excel_path)
        ws = wb['ran']
        for row in range(2, ws.max_row + 1):
            if ws['E' + str(row)].value != 0 and ws['F' + str(row)].value == 'OUTCODE':
                outcode_list.append( ws['G' + str(row)].value)
        return outcode_list

    def start_requests(self):
        url_prefix = 'https://www.rightmove.co.uk/property-to-rent/find.html?locationIdentifier=OUTCODE%5E'
        outcode_list = self.extract_code_from_excel()
        
        for outcode in outcode_list:
            url = url_prefix + str(outcode)
            yield scrapy.Request(url=url, callback=self.parse)
    
    def parse(self, response):
        self.logger.info('Parse function called on %s', response.url)
        Extract_total = response.xpath('//span[@class="searchHeader-resultCount"]/text()').get()
        Extract_total = Extract_total.replace(',','')
        This_area_total = int(Extract_total)

        parsed = urlparse(response.url)
        querys = parse_qs(parsed.query)
        para = {k: v[0] for k, v in querys.items()}
        para['index'] = int(para['index']) + 24 if 'index' in para else 24
        if para['index'] < This_area_total:
            # print("地区：%s; 总房源数: %s；目前第%s页；准备下一页爬虫" % (para['locationIdentifier'],This_area_total, para['index']//24))
            data = urlencode(para)
            next_page = 'https://www.rightmove.co.uk/property-to-rent/find.html?' + data
            yield response.follow(next_page, callback=self.parse)
            
        htmls = response.xpath('//a[contains(@href,"/properties")]/@href').getall()
        for i in htmls:
            if i not in self.seen:
                self.seen.add(i)
                yield response.follow(i, callback=self.parse_details)
        for k,v in self.try_again.items():
            if v >= 5 :
                del self.try_again[k]
                self.logger.info('%s won\'t parse again because failed to extract  %s times'% (k,v))
            else:
                yield response.follow(k, callback=self.parse_details)

    def parse_details(self, response):
        self.count += 1
        item = ScrapyrightmoveItem()
        data = extract_details(response)
        if not data:
            self.try_again[response.url]+=1
            self.logger.warning("%s runs %s times extract_details func failed" %(response.url, self.try_again[response.url]))
            return
        try:
            item['id'] = extract_id(data)
            item['title'] = extract_title(data)
            item['address'] = extract_address(data)
            item['agent'] = extract_agentdetails(data)
            item['rent'] = extract_rent(data)
            item['latitude'], item['longitude'] = extract_location(data)
            item['postcode'] = extract_postcode(data)
            # due to original url are filtered by letagree=False, so don't need to extract
            # letagreed = extract_letagreed(data)
            item['create_date'] = extract_addtime(data)
            item['url'] = response.url
            item['let_agreed'] = extract_letagreed(data)
            item['thumbnail'] = extract_image(data)
            yield item
        except Exception as e:
            self.logger.error("parse_details error: %s" % str(e))
            print("extraction error:", response.url)